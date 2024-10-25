import time
import win32com.client
import re
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import sys

def obtener_nombre_archivo(carpeta):
    archivos = os.listdir(carpeta)
    archivos_desvinculados = [archivo for archivo in archivos if archivo.startswith("Desvinculados") and archivo.endswith(".xlsx")]
    
    if archivos_desvinculados:
        nombre_archivo_antiguo = archivos_desvinculados[0]
        return nombre_archivo_antiguo
    else:
        print("No se encontró ningún archivo que comience con 'Desvinculados'.")
        return None

def leer_excel(carpeta):
    archivos = os.listdir(carpeta)
    archivos_desvinculados = [archivo for archivo in archivos if archivo.startswith("Desvinculados") and archivo.endswith(".xlsx")]
    
    if archivos_desvinculados:
        ruta_archivo = os.path.join(carpeta, archivos_desvinculados[0])
        datos = pd.read_excel(ruta_archivo)
        return datos
    else:
        print("No se encontró ningún archivo que comience con 'Desvinculados'.")
        return None

def Comprobar_Datos(cuerpo_mensaje):
    validador1 = cuerpo_mensaje.find("han dejado de pertenecer a Mutual")
    validador2 = cuerpo_mensaje.find("ha determinado dejar sin efecto el término de la relación laboral")

    if validador1 != -1:
        print("Nuevos desvinculados: \n")
        time.sleep(1)
        return 1    
    elif validador2 != -1:
        print("Reincorporación laboral: \n")
        return 2
    else:
        print("Correo sin información sobre desvinculados: \n")
        return 3

def Leer_Mensaje(remitente):
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    bandeja_entrada = outlook.GetDefaultFolder(6)

    subcarpeta_desvinculados = bandeja_entrada.Folders["Desvinculados Nuevos"]
    mensajes = subcarpeta_desvinculados.Items
    mensajes_filtrados = mensajes.Restrict(f"[SenderName] = '{remitente}'")
    mensajes_filtrados.Sort("[ReceivedTime]", True)
    ultimo_mensaje = mensajes_filtrados.GetLast()

    if ultimo_mensaje:
        cuerpo_mensaje = ultimo_mensaje.Body
        cuerpo_mensaje_sin_puntos = cuerpo_mensaje.replace(".", "")
        return cuerpo_mensaje_sin_puntos, ultimo_mensaje
    else:
        return None, None

def Extraer_Datos(cuerpo_mensaje):
    if cuerpo_mensaje is None:
        return "El cuerpo del correo es None."

    inicio = cuerpo_mensaje.find("SALDOS PENDIENTES")
    if inicio == -1:
        return "No se encontró la sección 'SALDOS PENDIENTES'"
    
    fin = cuerpo_mensaje.find('“', inicio)
    if fin == -1:
        fin = len(cuerpo_mensaje)
    
    return cuerpo_mensaje[inicio:fin].strip()

def Extraer_Solo_Nombres(datos_extraidos):
    palabras_excluir = ["K", "k", "SALDOS PENDIENTES", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE", "RUT", "NOMBRE", "FECHA DE RETIRO"]
    texto_limpio = re.sub(r'[0-9-]', '', datos_extraidos)

    for palabra in palabras_excluir:
        texto_limpio = re.sub(rf'\b{re.escape(palabra)}\b', '', texto_limpio)

    return texto_limpio

def Capturar_Nombres(texto):
    lineas = [line.strip() for line in texto.split('\n') if line.strip()]
    return lineas

def Capturar_RUTs(texto):
    patron_rut = r"\b\d{1,2}\.?\d{3}\.?\d{3}-[0-9Kk]\b|\b\d{7,8}-[0-9Kk]\b"
    ruts = re.findall(patron_rut, texto)
    return ruts

def Capturar_Fechas(texto):
    patron_fecha = r'\b\d{2}-\d{2}-\d{4}\b'
    fechas = re.findall(patron_fecha, texto)
    return fechas

def Crear_DataFrame(ruts, nombres, fechas):
    df = pd.DataFrame({
        'RUT': ruts,
        'NOMBRE': nombres,
        'FECHA DE RETIRO': fechas
    })
    
    return df

def Quitar_Duplicados(dataframe_combinado):
    dataframe_combinado['combinada'] = dataframe_combinado.iloc[:, 0].astype(str) + dataframe_combinado.iloc[:, 2].astype(str)
    
    dataframe_final = dataframe_combinado.drop_duplicates(subset='combinada')

    dataframe_final = dataframe_final.drop(columns=['combinada'])
    
    return dataframe_final

def Crear_Excel_Nuevo(dataframe_final):
    hoy = date.today()
    ruta_nuevo_excel = r'C:\Users\agalindo\OneDrive - Mutual\Escritorio\Pro. Desvinculados\DESV Y PLANT\Desvinculados - ' + hoy.strftime('%Y-%m-%d') + '.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    for r_idx, row in enumerate(dataframe_to_rows(dataframe_final, index=False, header=True), 1):
        ws.append(row)

    ws.column_dimensions[get_column_letter(1)].width = 15  # Columna 1
    ws.column_dimensions[get_column_letter(2)].width = 45  # Columna 2
    ws.column_dimensions[get_column_letter(3)].width = 20  # Columna 3

    date_style = NamedStyle(name="date_style", number_format='DD/MM/YYYY')

    for cell in ws[get_column_letter(3)]:
        cell.style = date_style

    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    wb.save(ruta_nuevo_excel)

def Borrar_Excel_Anterior(carpeta, nombre_archivo_antiguo):
    ruta_archivo_antiguo = os.path.join(carpeta, nombre_archivo_antiguo)

    if os.path.exists(ruta_archivo_antiguo):
        try:
            os.remove(ruta_archivo_antiguo) 
            print(f"Archivo anterior '{nombre_archivo_antiguo}' eliminado con éxito.")
        except Exception as e:
            print(f"No se pudo eliminar el archivo anterior: {e}")
    else:
        print(f"El archivo '{nombre_archivo_antiguo}' no existe.")

def Mover_Mensaje(remitente, mensaje):
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    bandeja_entrada = outlook.GetDefaultFolder(6)

    subcarpeta_destino = bandeja_entrada.Folders["Desvinculados Listos"]
    mensaje.Move(subcarpeta_destino)
    print("\nOUTLOOK: El mensaje se ha movido a 'Desvinculados Listos'.")

def Mover_Mensaje_error(remitente, mensaje):
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    bandeja_entrada = outlook.GetDefaultFolder(6)

    subcarpeta_destino = bandeja_entrada.Folders["Desvinculados Error"]
    mensaje.Move(subcarpeta_destino)
    print("El mensaje se ha movido a 'Desvinculados erróneos'.")

carpeta = r'C:\Users\agalindo\OneDrive - Mutual\Escritorio\Pro. Desvinculados\DESV Y PLANT'
remitente = "Aramis Ignacio Galindo Vidal"
excel_seleccionado = leer_excel(carpeta)
nombre_archivo_antiguo = obtener_nombre_archivo(carpeta)
df_acumulado = pd.DataFrame(columns=['RUT', 'NOMBRE', 'FECHA DE RETIRO'])
validador = 0

dataframes = []

while True:
    mensaje, ultimo_mensaje = Leer_Mensaje(remitente)
    
    if not mensaje:
        print(f"No se encontraron más mensajes de {remitente}.")
        break

    contenido_mensaje = Comprobar_Datos(mensaje)
    
    if contenido_mensaje == 1:
        datos_extraidos = Extraer_Datos(mensaje)
        solo_nombres = Extraer_Solo_Nombres(datos_extraidos)

        nombres = Capturar_Nombres(solo_nombres)
        ruts = Capturar_RUTs(datos_extraidos)
        fechas = Capturar_Fechas(datos_extraidos)

        tamaño_rut = len(ruts)
        tamaño_fechas = len(fechas)
        tamaño_nombres = len(nombres)

        if tamaño_rut == tamaño_fechas == tamaño_nombres:
            df = Crear_DataFrame(ruts, nombres, fechas)
            print("Nuevo DataFrame capturado: \n", df)
            
            dataframes.append(df)
            
            Mover_Mensaje(remitente, ultimo_mensaje)                

        else:
            print("Error, tamaños desiguales")
            print(tamaño_rut)
            print(tamaño_fechas)
            print(tamaño_nombres)

    elif contenido_mensaje == 2:
        Mover_Mensaje(remitente, ultimo_mensaje)

    elif contenido_mensaje == 3:
        Mover_Mensaje_error(remitente, ultimo_mensaje)

if dataframes:
    df_final = pd.concat(dataframes, ignore_index=True)
    print("DataFrame combinado de todos los mensajes: \n", df_final)

    if excel_seleccionado is not None:
        dataframe_combinado = pd.concat([excel_seleccionado, df_final], ignore_index=True)
        dataframe_final = Quitar_Duplicados(dataframe_combinado)

        Crear_Excel_Nuevo(dataframe_final)

        hoy = date.today()
        nombre_archivo_nuevo = 'Desvinculados - ' + hoy.strftime('%Y-%m-%d') + '.xlsx'
        
        print("")

        if nombre_archivo_antiguo != nombre_archivo_nuevo:
            Borrar_Excel_Anterior(carpeta, nombre_archivo_antiguo)
            print("Nuevo documento creado con éxito\n")
        
        validador = 1

else:
    print("No se encontraron nuevos registros de desvinculados.")

if validador == 1:
    print("Fin de extracción de datos, armando EXCEL")

    dataframe_nuevo = leer_excel(carpeta)
    dataframe_nuevo['FECHA DE RETIRO'] = pd.to_datetime(dataframe_nuevo['FECHA DE RETIRO'], errors='coerce').dt.strftime('%Y/%m/%d')
    dataframe_nuevo['PERIODO'] = pd.to_datetime(dataframe_nuevo['FECHA DE RETIRO'], errors='coerce').dt.strftime('%Y%m')

    print(dataframe_nuevo)
    Crear_Excel_Nuevo(dataframe_nuevo)

print("Proceso terminado con éxito")